# -*- coding:utf-8 -*-
# !/usr/bin/env python
#---------------------------------------------------------------------------------
# Author: Zhang
#
# Create Date: 2024/11/14
# Last Update on: 2025/01/06
#
# FILE: functions.py
# Description: Basic functions are defined here
#---------------------------------------------------------------------------------

#---------------------------------------------------------------------------------
# IMPORT REQUIRED PACKAGES HERE

import os
import sys
import math
import time
import platform
import openpyxl

from pathlib import Path

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

import utils.components as comps
import utils.instrument as instr

# END OF PACKAGE IMPORT
#---------------------------------------------------------------------------------

#---------------------------------------------------------------------------------
# DEFINE GLOBAL VARIABLES HERE

t_cal = comps.TraceDataCalculator()

comps.Const.MAIN_DIR = Path(__file__).resolve().parent.parent
comps.Const.SSHOT_DIR = os.path.join(comps.Const.MAIN_DIR, 'screenshots')
comps.Const.TRACE_DIR = os.path.join(comps.Const.MAIN_DIR, 'trace_data')

# END OF GLOBAL VARIABLES DEFINITION
#---------------------------------------------------------------------------------

#---------------------------------------------------------------------------------
# DEFINE FUNCTIONS HERE

def menu() -> None:
    print(f"- Initialize successfully. Starting the program...")
    time.sleep(2)
    os.system('cls') if platform.system() == 'Windows' else os.system('clear')
    print(f'# ----------------------------------------------------------------------------- #')
    print(f'# GITEKI PRE-TEST AUTOMATION TOOL')
    print(f'# INTERNAL USE ONYL @ RENESAS/A&C/CONN/LIC/ACENG')
    print(f'# AUTHOR: ZHANG')
    print(f'# CREATED ON: 2024/11/14')
    print(f'# For now (2024/11/28), all output about power is reading value, not EIRP.')
    print(f'# ----------------------------- COMMAND LIST ---------------------------------- #')
    print(f'# \"obw\" or \"sbw\"  :     Start the measurement of OBW and SBW.')
    print(f'# \"peak power\"    :     Start the measurement of peak power.')
    print(f'# \"ave power\"     :     Start the measurement of average power.')
    print(f'# \"spurious\"      :     Start the measurement of peak and averagr spurious.')
    print(f'# \"plot\"          :     Plot trace data and mask. (Mannully save needed)')
    print(f'# \"set rule\"      :     Re-select the rule (49_27_3 or 49_27_4).')
    print(f'# \"exit\"          :     Exit the program.')
    print(f'# ----------------------------------------------------------------------------- #')
    print(f'')

def show_measurement_result(result_dict: dict, headers=None):
    def format_row(*args):
        return "| " + " | ".join(f"{str(arg):<{col_widths[i]}}" for i, arg in enumerate(args)) + " |"
    
    if headers is None:
        headers = ["Item", "Value", "Passed/Failed"]
    col_widths = [max(len(str(key)) for key in result_dict.keys())]
    for i in range(1, len(headers)):
        col_widths.append(max(
            len(str(value[i - 1])) if isinstance(value, list) else len(str(value))
            for value in result_dict.values()
        ))
    col_widths = [max(width, len(headers[i])) for i, width in enumerate(col_widths)]
    separator = "+" + "+".join("-" * (width + 2) for width in col_widths) + "+"

    print(f"{separator}\n{format_row(*headers)}\n{separator}")
    for key, value in result_dict.items():
        row = [key] + (value if isinstance(value, list) else [value, "-----"])
        print(format_row(*row[:len(headers)]))
    print(separator)

def show_configure(config_dict: dict):
    def format_row(items):
        return "| " + " | ".join(f"{str(item):<{col_widths[i]}}" for i, item in enumerate(items)) + " |"
    
    keys, values = list(config_dict.keys()), list(config_dict.values())
    col_widths = [max(len(str(key)), len(str(value))) for key, value in config_dict.items()]
    separator = "+" + "+".join("-" * (width + 2) for width in col_widths) + "+"

    print(f"- Measureing Condition: ")
    print(f"{separator}\n{format_row(keys)}\n{separator}\n{format_row(values)}\n{separator}")

def write_report(report_file, text, data_dict):
    try:
        workbook = openpyxl.load_workbook(report_file)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
    sheet = workbook.active

    row = 1
    while sheet.cell(row=row, column=1).value is not None:
        row += 1

    sheet.cell(row=row, column=1, value=text)
    row += 1

    for key, value in data_dict.items():
        sheet.cell(row=row, column=1, value=key)
        if isinstance(value, list) and len(value) == 2:
            sheet.cell(row=row, column=2, value=value[0])
            sheet.cell(row=row, column=3, value=value[1])
        else:
            sheet.cell(row=row, column=2, value=value)
        row += 1

    workbook.save(report_file)
    workbook.close()

def choose_condition(condition_type, prompt=None) -> str:
    if condition_type == 'rule':
        print("- Please select a rule:  (1) 49_27_3;  (2) 49_27_4.")
        user_input = input("Waiting for index of rule: ")
        rule = "49_27_3" if user_input == '1' else "49_27_4"
        print(f"#--------------------------------------------------------------#")
        print(f"#  All measurements will be conducted to align rule: {rule}.  #")
        print(f"#--------------------------------------------------------------#\n")
        return rule
    
    elif condition_type == 'method':
        print("- Please select a method:  (1) general;  (2) exception.")
        print(f"  - General  : Basic method describled in rule book.")
        print(f"  - Exception: {prompt}")
        user_input = input("Waiting for index of method: ")
        return "general" if user_input == '1' else "exception"
    else:
        return None

def unit_measurement(sa: instr.SpectrumAnalyzer, sa_config: dict):
    show_configure(sa_config)  # display spectrum analyzer configure

    sa.config(param='continues_sweep', value='OFF')  # stop sweep
    input('- Waiting for board launch up. Press enter to start measurement.')

    sa.config(param=sa_config)  # start continues sweep
    input(f"- Waiting trace to stabilize. Press Enter to stop sweep.")
    sa.config(param='continues_sweep', value='OFF')  # stop sweep

    peak_freq, peak_power = sa.aquire_peak_point()
    return peak_freq, peak_power

def measure_obw_and_sbw(sa: instr.SpectrumAnalyzer, giteki_dict: dict, rule='49_27_3'):
    # calculate center freq. and span
    start_freq = float(giteki_dict['masks'][rule]['obw'].split('~')[0])  #GHz
    stop_freq = float(giteki_dict['masks'][rule]['obw'].split('~')[1])   #GHz
    sbw_limit = float(giteki_dict['masks']['sbw']) * 1e3  # MHz
    obw_limit = (stop_freq - start_freq) * 1e3  # MHz
    center_freq = (start_freq + stop_freq) / 2  # GHz
    
    # place tmp marker for obw measurement
    sa.config(param='obw_percent', value='99PCT')
    sa.config(param='obw_measure', value='SEL OBW', delimiter=':')
    sa.config(param='obw_measure', value='RES? OBW', delimiter=':')

    # construct other cfg of spectrum analyzer
    sa_cfg = {'center_freq': f"{center_freq}GHz", 'span': f"{obw_limit * 3}MHz"}
    sa_cfg.update(giteki_dict['obw_and_sbw'])

    unit_measurement(sa, sa_cfg)  # perform measurement
    sa.save_screenshot(os.path.join(comps.Const.SSHOT_DIR, f'obw_and_sbw.png'))  # save screenshot
    sa.save_trace_to_csv(os.path.join(comps.Const.TRACE_DIR, f'obw_and_sbw.csv'))  # save trace data

    # use trace data to calculate obw and sbw
    global t_cal
    t_cal.read_trace_data('obw_and_sbw.csv')
    obw, l_freq, r_freq = t_cal.calculate_obw()  # calculate obw
    obw_status = "Passed" if obw <= obw_limit else "Failed"
    l_freq_status = "Passed" if l_freq >= start_freq else "Failed"
    r_freq_status = "Passed" if r_freq <= stop_freq else "Failed"
    sbw = t_cal.calculate_sbw()  # calculate abw
    sbw_status = "Passed" if sbw >= sbw_limit else "Failed"

    return {
        'obw': [f"{obw}MHz", obw_status],
        'obw lower freq': [f"{l_freq}GHz", l_freq_status],
        'obw upper freq': [f"{r_freq}GHz", r_freq_status],
        'sbw': [f"{sbw}MHz", sbw_status]
    }

def measure_peak_power(sa: instr.SpectrumAnalyzer, giteki_dict: dict, rule='49_27_3', method='general'):
    start_freq = f"{giteki_dict['masks'][rule]['obw'].split('~')[0]}GHz"
    stop_freq = f"{giteki_dict['masks'][rule]['obw'].split('~')[1]}GHz"

    # construct spectrum analyzer config
    sa_cfg = {'start_freq': start_freq, 'stop_freq': stop_freq}
    sa_cfg.update(giteki_dict['search']['common'])
    sa_cfg.update(giteki_dict['search']['diff']['peak']['general'])  # set RBW to 3MHz, VBW to 10MHz

    search_freq, _ = unit_measurement(sa, sa_cfg)  # search
    sa.save_screenshot(os.path.join(comps.Const.SSHOT_DIR, f'peak_search.png'))

    # zoom in with span 100MHz
    del sa_cfg['start_freq']
    del sa_cfg['stop_freq']
    sa_cfg.update({'center_freq': search_freq, 'span': '100MHz'})
    if method != 'general':  # wide RBW to 50MHz if necessary
        sa_cfg.update(giteki_dict['search']['diff']['peak']['exception'])
    
    search_freq, search_peak = unit_measurement(sa, sa_cfg)
    sa.save_screenshot(os.path.join(comps.Const.SSHOT_DIR, f'peak_zoom_in_{method}.png'))

    # calculate correction factor
    rbw = float(sa_cfg['RBW'].split('MHz'))
    corr_factor = 20 * math.log10(float(50.0) / rbw)
    peak = corr_factor + search_peak

    # load limit
    obw_key = giteki_dict['masks'][rule]['obw']
    limit = giteki_dict['masks'][rule][obw_key]['peak']
    dev = giteki_dict['masks'][rule][obw_key]['allowable_dev']
    limit = 10 * math.log10((10 ** (limit / 10)) * (1 + dev))

    status = 'Passed' if peak <= limit else 'Failed'

    return {
        'reading value': f"{round(search_peak, 2)}dBm",
        'freq.': f"{round(search_freq / 1e9, 5)}GHz",
        'corr. factor': f"{round(corr_factor, 2)}",
        'peak (reading val. + corr. factor)': [f"{round(peak, 2)}dBm", status]
    }

def measure_ave_power(sa: instr.SpectrumAnalyzer, giteki_dict: dict, rule='49_27_3', method='general'):
    freq_interval = giteki_dict['masks'][rule]['obw']
    start_freq = f"{freq_interval.split('~')[0]}GHz"
    stop_freq = f"{freq_interval.split('~')[1]}GHz"

    # construct spectrum analyzer config
    sa_cfg = {'start_freq': start_freq, 'stop_freq': stop_freq}
    sa_cfg.update(giteki_dict['search']['common'])

    search_freq, _ = unit_measurement(sa, sa_cfg)  # search
    sa.save_screenshot(os.path.join(comps.Const.SSHOT_DIR, f'ave_search.png'))
    sa.save_trace_to_csv(os.path.join(comps.Const.TRACE_DIR, f'trace_{freq_interval}.csv'))  # to draw plot

    # zoom in with span 100MHz and 10MHz
    del sa_cfg['start_freq']
    del sa_cfg['stop_freq']
    for span in ['100MHz', '10MHz']:
        sa_cfg.update({'center_freq': search_freq, 'span': span})
        search_freq, _ = unit_measurement(sa, sa_cfg)
        sa.save_screenshot(os.path.join(comps.Const.SSHOT_DIR, f'ave_zoom_in_span={span}.png'))

    # measure step
    sa_cfg.update({'center_freq': search_freq})
    if method == 'general':  # use SMP detector to measure and claculate
        sa_cfg.update(giteki_dict['measure']['common'])
        measure_freq, _ = unit_measurement(sa, sa_cfg)
        sa.save_trace_to_csv(os.path.join(comps.Const.TRACE_DIR, f'ave_measure.csv'))

        # use trace data to calculate average power
        global t_cal
        t_cal.read_trace_data(os.path.join(comps.Const.TRACE_DIR, f'ave_measure.csv'))
        ave = t_cal.calculate_ave_power()

    else:  # use RMS detector to measure
        sa_cfg.update(giteki_dict['measure']['diff']['ave'])
        measure_freq, ave = unit_measurement(sa, sa_cfg)

    sa.save_screenshot(os.path.join(comps.Const.SSHOT_DIR, f'ave_measure_{method}.png'))
    
    # load limit
    obw_key = giteki_dict['masks'][rule]['obw']
    d_freq = giteki_dict['masks'][rule][obw_key]['dividing_freq']
    idx = '1' if measure_freq <= d_freq else '2'
    limit = giteki_dict['masks'][rule][obw_key][f'ave{idx}']
    dev = giteki_dict['masks'][rule][obw_key]['allowable_dev']
    limit = 10 * math.log10((10 ** (limit / 10)) * (1 + dev))

    status = 'Passed' if ave <= limit else 'Failed'

    return {
        'ave power': [f"{round(ave, 2)}dBm", status],
        'freq': f"{round(measure_freq / 1e9, 5)}GHz"
    }

def measure_spurious(sa: instr.SpectrumAnalyzer, giteki_dict: dict, rule='49_27_3'):
    result = {}
    freq_intervals = giteki_dict['masks'][rule].copy()
    
    del freq_intervals[giteki_dict['masks'][rule]['obw']]
    del freq_intervals['obw']

    # iterate all spurious freq. band
    for f_interval in freq_intervals:
        start_freq = f"{f_interval.split('~')[0]}GHz"
        stop_freq = f"{f_interval.split('~')[1]}GHz"

        # search step
        sa_cfg = {'start_freq': start_freq, 'stop_freq': stop_freq}
        sa_cfg.update(giteki_dict['search']['common'])

        search_freq, search_peak = unit_measurement(sa, sa_cfg)
        sa.save_screenshot(os.path.join(comps.Const.SSHOT_DIR, f'spu_search_{f_interval}.png'))
        # to draw plot
        sa.save_trace_to_csv(os.path.join(comps.Const.TRACE_DIR, f'trace_{f_interval}.csv'))  

        ave_limit = giteki_dict['masks'][rule][f_interval]['ave']
        peak_limit = giteki_dict['masks'][rule][f_interval]['peak']

        peak = ave = cal_ave = peak_freq = ave_freq = None
        if search_peak <= ave_limit:
            ave, ave_freq = search_peak, search_freq

        if search_peak <= peak_limit - 3:
            peak, peak_freq = search_peak, search_freq

        if peak is None or ave is None:  # zoom in step if necessary
            del sa_cfg['start_freq']
            del sa_cfg['stop_freq']

            for span in ['100MHz', '10MHz']:  # zoom in step
                sa_cfg.update({'center_freq': search_freq, 'span': span})
                search_freq, search_peak = unit_measurement(sa, sa_cfg)
                sa.save_screenshot(os.path.join(comps.Const.SSHOT_DIR, f'spu_zoom_in_span={span}_{f_interval}.png'))

            if peak is None:  # update peak result if necessary
                peak, peak_freq = search_peak, search_freq
            
            if ave is None:  # measure step if necessary
                sa_cfg.update({'center_freq': search_freq})
                sa_cfg.update(giteki_dict['measure']['common'])
                ave_freq, ave = unit_measurement(sa, sa_cfg)

                # save trace data to get calculate average
                sa.save_trace_to_csv(os.path.join(comps.Const.TRACE_DIR, f'spu_ave_measure_{f_interval}.csv'))
                global t_cal
                t_cal.read_trace_data(os.path.join(comps.Const.TRACE_DIR, f'spu_ave_measure_{f_interval}.csv'))
                cal_ave = t_cal.calculate_ave_spurious()

        # pass / fail determination
        peak_status = 'Passed' if peak <= peak_limit else 'Failed'
        m_ave_status = 'Passed' if ave <= ave_limit else 'Failed'
        if cal_ave:
            c_ave_status = 'Passed' if cal_ave <= ave_limit else 'Failed'
        else:
            c_ave_status = 'None'

        result[f_interval] = {
            'spurious peak': [f'{round(peak)}dBm', peak_status],
            'peak freq': f'{round(peak_freq) / 1e9}GHz',
            'spurious ave (reading val)': [f'{round(ave)}dBm', m_ave_status],
            'spurious ave (cal.)': [f'{round(cal_ave)}dBm', c_ave_status],
            'ave freq.': f'{round(ave_freq) / 1e9}GHz'
        }

    return result